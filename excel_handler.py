import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO


class ExcelManager:
    def __init__(self, template_path):
        self.wb = openpyxl.load_workbook(template_path)

    def get_sheet_names(self):
        return self.wb.sheetnames

    def create_promo_tab(self, base_sheet_name, new_tab_name):
        base_sheet = self.wb[base_sheet_name]
        new_sheet = self.wb.copy_worksheet(base_sheet)
        new_sheet.title = new_tab_name
        return new_sheet

    def write_to_cell(self, sheet_name, cell_ref, value):
        self.wb[sheet_name][cell_ref] = value

    def read_cell(self, sheet_name, cell_ref):
        return self.wb[sheet_name][cell_ref].value

    def append_dataframe(self, sheet_name, df):
        sheet = self.wb[sheet_name]
        max_row = sheet.max_row

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            if max_row > 1 and r_idx == 1:
                continue
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=max_row + r_idx - 1, column=c_idx, value=value)

    def write_vertical_array(self, sheet_name, start_cell, data_list):
        sheet = self.wb[sheet_name]
        col = start_cell[0]
        start_row = int(start_cell[1:])
        for i, val in enumerate(data_list):
            try:
                val_to_write = float(val)
            except ValueError:
                val_to_write = val
            sheet[f"{col}{start_row + i}"] = val_to_write

    def get_download_bytes(self):
        output = BytesIO()
        self.wb.save(output)
        output.seek(0)
        return output
    
    def read_column(self, sheet_name, column_letter):
        try:
            ws = self.wb[sheet_name]
            # Get all cells in the specified column
            column_cells = ws[column_letter]
            
            # Extract text from cells, ignoring empty ones
            sql_lines = [str(cell.value) for cell in column_cells if cell.value is not None]
            
            return "\n".join(sql_lines)
        except Exception as e:
            return f"-- Error reading SQL from {sheet_name} column {column_letter}: {e}"
        
    def overwrite_item_list(self, sheet_name, df):
        import pandas as pd
        from openpyxl.utils.dataframe import dataframe_to_rows
        
        # This guarantees we only touch the specific sheet (item-List)
        sheet = self.wb[sheet_name]
        
        # 1. Wipe ONLY this specific sheet clean
        sheet.delete_rows(1, sheet.max_row)
        
        # 2. Paste the uploaded data exactly as-is, starting at A1
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), 1):
            for c_idx, value in enumerate(row, 1):
                # Skip blank cells so we don't paste "NaN"
                if pd.notna(value):
                    sheet.cell(row=r_idx, column=c_idx, value=value)

    def remove_unwanted_sheets(self, sheets_to_remove):
        """Deletes specified backend/template sheets before final export."""
        for sheet in sheets_to_remove:
            if sheet in self.wb.sheetnames:
                del self.wb[sheet]

    def add_raw_sheet(self, sheet_name, df):
        """Creates a new sheet and pastes a dataframe into it exactly as-is."""
        if sheet_name in self.wb.sheetnames:
            del self.wb[sheet_name]
        ws = self.wb.create_sheet(sheet_name)
        
        from openpyxl.utils.dataframe import dataframe_to_rows
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

    def write_kv_pairs(self, sheet_name, data_dict, mapping_dict):
        """
        Writes values to specific cells based on a key-mapping.
        :param sheet_name: Target worksheet name
        :param data_dict: Parsed SQL data { 'Metric_A': 100, 'Metric_B': 200 }
        :param mapping_dict: Mapping from JSON { 'Metric_A': 'B10', 'Metric_B': 'C10' }
        """
        sheet = self.wb[sheet_name]
        for key, value in data_dict.items():
            if key in mapping_dict:
                cell_address = mapping_dict[key]
                try:
                    # Try to convert to float for Excel calculations
                    val_to_write = float(value)
                except (ValueError, TypeError):
                    val_to_write = value
                
                sheet[cell_address] = val_to_write
            else:
                # Optional: Log or print keys that aren't mapped
                print(f"Warning: Key '{key}' not found in mapping for sheet {sheet_name}")